# -*- coding: utf-8 -*-
"""
Created on Thu Mar 12 23:35:56 2020

@author: ASUS
"""

import  xdrlib ,sys
import xlrd
import pandas as pd

import xlwt
import xlrd
from datetime import date,datetime

import numpy as np


from openpyxl import load_workbook

#file = 'G:\\工作\\联通\\数据\\20200325\\3月24日分公司小程序明细.xlsx'

#设置表格样式
def set_style(name,height,bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    return style

"""
def read_excel():

  
    wb = xlrd.open_workbook(filename=file)#打开文件
    print(wb.sheet_names())#获取所有表格名字

    sheet1 = wb.sheet_by_index(0)#通过索引获取表格
    sheet2 = wb.sheet_by_name('Sheet2')#通过名字获取表格
    print(sheet1,sheet2)
    print(sheet1.name,sheet1.nrows,sheet1.ncols)

    rows = sheet1.row_values(2)#获取行内容
    cols = sheet1.col_values(3)#获取列内容
    print(rows)
    print(cols)

    print(sheet1.cell(2,0).value)#获取表格里的内容，三种方式
    print(sheet1.cell_value(1,0))
    print(sheet1.row(1)[0].value)
    
    
    
#写Excel
def write_excel():
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('学生',cell_overwrite_ok=True)
    row0 = ["姓名","年龄","出生日期","爱好"]
    colum0 = ["张三","李四","恋习Python","小明","小红","无名"]
    #写第一行
    for i in range(0,len(row0)):
        sheet1.write(0,i,row0[i],set_style('Times New Roman',220,True))
    #写第一列
    for i in range(0,len(colum0)):
        sheet1.write(i+1,0,colum0[i],set_style('Times New Roman',220,True))

    sheet1.write(1,3,'2006/12/12')
    sheet1.write_merge(6,6,1,3,'未知')#合并行单元格
    sheet1.write_merge(1,2,3,3,'打游戏')#合并列单元格
    sheet1.write_merge(4,5,3,3,'打篮球')

    f.save('test2.xls')    
    
"""
    
def gather_excel(file,file_reasoult):

    wb = xlrd.open_workbook(filename=file)#打开文件
    print(wb.sheet_names())#获取所有表格名字
    
    #通过名字获取表格
    sheet_YWWL = wb.sheet_by_name('移网本地')
    sheet_YWXC = wb.sheet_by_name('移网下沉')
    sheet_LLB = wb.sheet_by_name('流量包')
    sheet_DK = wb.sheet_by_name('单宽')
    sheet_RH = wb.sheet_by_name('融合')    
 
    
    print(sheet_YWWL.name,sheet_YWWL.nrows,sheet_YWWL.ncols)
    #获取行内容
    rows1 = sheet_YWWL.row_values(0)
    print(rows1)   
    #获取列内容 从0开始的
    cols_QDBM_YWWL = sheet_YWWL.col_values(30) #渠道编码
    cols_QDMC_YWWL = sheet_YWWL.col_values(31) #渠道名称
    cols_FZRBM_YWWL = sheet_YWWL.col_values(32)#发展人编码
    cols_FZRXM_YWWL = sheet_YWWL.col_values(33)#发展人姓名
    cols_XDFGS_YWWL = sheet_YWWL.col_values(34) #下单分公司
    cols_YWMC_YWWL = sheet_YWWL.col_values(10) #业务名称
    
    
    print(sheet_YWXC.name,sheet_YWXC.nrows,sheet_YWXC.ncols)
    #获取行内容
    rows2 = sheet_YWXC.row_values(0)
    print(rows2)   
    #获取列内容 从0开始的
    cols_QDBM_YWXC = sheet_YWXC.col_values(38) #渠道编码
    del cols_QDBM_YWXC[0] #去除第一行
    cols_QDMC_YWXC = sheet_YWXC.col_values(39) #渠道名称
    del cols_QDMC_YWXC[0] #去除第一行
    cols_FZRBM_YWXC = sheet_YWXC.col_values(40)#发展人编码
    del cols_FZRBM_YWXC[0] #去除第一行
    cols_FZRXM_YWXC = sheet_YWXC.col_values(41)#发展人姓名
    del cols_FZRXM_YWXC[0] #去除第一行
    cols_XDFGS_YWXC = sheet_YWXC.col_values(42) #下单分公司
    del cols_XDFGS_YWXC[0] #去除第一行
    cols_YWMC_YWXC = sheet_YWXC.col_values(3) #业务名称
    del cols_YWMC_YWXC[0] #去除第一行    
    
    
    print(sheet_LLB.name,sheet_LLB.nrows,sheet_LLB.ncols)
    #获取行内容
    rows3 = sheet_LLB.row_values(0)
    print(rows3)   
    #获取列内容 从0开始的
    cols_QDBM_LLB = sheet_LLB.col_values(30) #渠道编码
    del cols_QDBM_LLB[0] #去除第一行
    cols_QDMC_LLB = sheet_LLB.col_values(31) #渠道名称
    del cols_QDMC_LLB[0] #去除第一行
    cols_FZRBM_LLB = sheet_LLB.col_values(32)#发展人编码
    del cols_FZRBM_LLB[0] #去除第一行
    cols_FZRXM_LLB = sheet_LLB.col_values(33)#发展人姓名
    del cols_FZRXM_LLB[0] #去除第一行
    cols_XDFGS_LLB = sheet_LLB.col_values(34) #下单分公司
    del cols_XDFGS_LLB[0] #去除第一行
    cols_YWMC_LLB = sheet_LLB.col_values(10) #业务名称
    del cols_YWMC_LLB[0] #去除第一行     
    
    
    print(sheet_DK.name,sheet_DK.nrows,sheet_DK.ncols)
    #获取行内容
    rows = sheet_DK.row_values(0)
    print(rows)   
    #获取列内容 从0开始的
    cols_QDBM_DK = sheet_DK.col_values(50) #渠道编码
    del cols_QDBM_DK[0] #去除第一行
    cols_QDMC_DK = sheet_DK.col_values(46) #渠道名称
    del cols_QDMC_DK[0] #去除第一行
    cols_FZRBM_DK = sheet_DK.col_values(48)#发展人编码
    del cols_FZRBM_DK[0] #去除第一行
    cols_FZRXM_DK = sheet_DK.col_values(47)#发展人姓名
    del cols_FZRXM_DK[0] #去除第一行
    cols_XDFGS_DK = sheet_DK.col_values(49) #下单分公司
    del cols_XDFGS_DK[0] #去除第一行
    cols_YWMC_DK = sheet_DK.col_values(5) #业务名称
    del cols_YWMC_DK[0] #去除第一行   
     
    print(sheet_RH.name,sheet_RH.nrows,sheet_RH.ncols)
    #获取行内容
    rows = sheet_RH.row_values(0)
    print(rows)   
    #获取列内容 从0开始的
    cols_QDBM_RH= sheet_RH.col_values(52) #渠道编码
    del cols_QDBM_RH[0] #去除第一行
    cols_QDMC_RH = sheet_RH.col_values(48) #渠道名称
    del cols_QDMC_RH[0] #去除第一行
    cols_FZRBM_RH = sheet_RH.col_values(50)#发展人编码
    del cols_FZRBM_RH[0] #去除第一行
    cols_FZRXM_RH = sheet_RH.col_values(49)#发展人姓名
    del cols_FZRXM_RH[0] #去除第一行
    cols_XDFGS_RH = sheet_RH.col_values(51) #下单分公司
    del cols_XDFGS_RH[0] #去除第一行    
    cols_YWMC_RH = sheet_RH.col_values(5) #业务名称
    del cols_YWMC_RH[0] #去除第一行       
      
    cols_QDBM = cols_QDBM_YWWL + cols_QDBM_YWXC + cols_QDBM_LLB + cols_QDBM_DK + cols_QDBM_RH
    cols_QDMC = cols_QDMC_YWWL + cols_QDMC_YWXC + cols_QDMC_LLB + cols_QDMC_DK + cols_QDMC_RH 
    cols_FZRBM = cols_FZRBM_YWWL + cols_FZRBM_YWXC + cols_FZRBM_LLB + cols_FZRBM_DK + cols_FZRBM_RH
    cols_FZRXM = cols_FZRXM_YWWL + cols_FZRXM_YWXC + cols_FZRXM_LLB + cols_FZRXM_DK + cols_FZRXM_RH
    cols_XDFGS = cols_XDFGS_YWWL + cols_XDFGS_YWXC + cols_XDFGS_LLB + cols_XDFGS_DK + cols_XDFGS_RH
    cols_YWMC = cols_YWMC_YWWL + cols_YWMC_YWXC + cols_YWMC_LLB + cols_YWMC_DK + cols_YWMC_RH
    
    
  #  lenth = len(cols_QDBM_YWWL)
#    print(cols_QDBM)
  #  print(lenth)
    
 
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('汇总',cell_overwrite_ok=False)  
    #写第一列
    for i in range(0,len(cols_QDBM)):
        sheet1.write(i,0,cols_QDBM[i])    
    #写第一列
    for i in range(0,len(cols_QDMC)):
        sheet1.write(i,1,cols_QDMC[i]) 
    #写第一列
    for i in range(0,len(cols_FZRBM)):
        sheet1.write(i,2,cols_FZRBM[i]) 
    #写第一列
    for i in range(0,len(cols_FZRXM)):
        sheet1.write(i,3,cols_FZRXM[i]) 
    #写第一列
    for i in range(0,len(cols_XDFGS)):
        sheet1.write(i,4,cols_XDFGS[i])         
    #写第一列
    for i in range(0,len(cols_YWMC)):
        sheet1.write(i,5,cols_YWMC[i])
       
    f.save(file_reasoult)   
    
    '''
    20200325 新增后续码表处理
    '''
    arr = np.array([cols_QDBM,cols_QDMC,cols_FZRBM,cols_FZRXM,cols_XDFGS,cols_YWMC])
    return arr.T

# file_read文件路径 sheet_index索引
def MyRead_excel(file_read,sheet_index):

  
    wb = xlrd.open_workbook(filename=file_read)#打开文件
    print(wb.sheet_names())#获取所有表格名字

    sheetX = wb.sheet_by_index(sheet_index)#通过索引获取表格
    #sheet2 = wb.sheet_by_name('Sheet2')#通过名字获取表格

    print(sheetX.name,sheetX.nrows,sheetX.ncols)
    
    arr = np.array(sheetX.row_values(0))
    for i in range(1,sheetX.nrows):
        rows = sheetX.row_values(i)#获取行内容
        arr = np.vstack((arr,rows))
    print(arr)
    return arr
    
# 20200325 重写COUNT函数
def MyExcel_COUNT(arr,col_index):


    
    
# 20200325 重写vlookup函数
    '''
    参数             简单说明                        输入数据类型
    lookup_value     要查找的值                      数值、引用或文本字符串
    table_array      要查找的区域                    数据表区域
    col_index_num    返回数据在查找区域的第几列数     正整数
    range_lookup     精确匹配/近似匹配               FALSE（或0）/TRUE（或1或不填）
    '''  
def MyExcel_VLOOKUP(Location,lookup_value,table_array,col_index_num,range_lookup):
    print('111')
    
    
def gather_excel_Dataframe(file_in,file_out):
    df_excel = pd.DataFrame(pd.read_excel(file_in))    
    

#20200325 新增后续码表处理

def gather_excel_DF(file,file_reasoult):
    df_sheet_YWWL = pd.DataFrame(pd.read_excel(file,sheet_name='移网本地'))
    df_sheet_YWXC = pd.DataFrame(pd.read_excel(file,sheet_name='移网下沉'))
    df_sheet_LLB = pd.DataFrame(pd.read_excel(file,sheet_name='流量包'))
    df_sheet_DK = pd.DataFrame(pd.read_excel(file,sheet_name='宽带'))
    #df_sheet_RH = pd.DataFrame(pd.read_excel(file,sheet_name='融合'))    
    #print(df_sheet_YWWL.columns)
    #print(df_sheet_YWXC.columns)
    #print(df_sheet_LLB.columns)
    #print(df_sheet_DK.columns)
    #print(df_sheet_RH.columns)    
    df_sheet_YWWL = df_sheet_YWWL[['系统来源','业务名称','订单时间','商品名称','渠道编码','渠道名称','发展人编码','发展人姓名','下单分公司']]
    df_sheet_YWXC = df_sheet_YWXC[['订单来源','业务类型','下单时间','商品名称','渠道编码','发展渠道名称','发展人编码','发展人名称','下单分公司']]   
    df_sheet_LLB = df_sheet_LLB[['系统来源','业务名称','订单时间','商品名称','渠道编码','渠道名称','发展人编码','发展人姓名','下单分公司']]
    df_sheet_DK = df_sheet_DK[['订单来源','业务类型','下单时间','商品名称','发展渠道编码','发展渠道名称','发展人编码','发展人名称','发展渠道所属分公司']]  
    #df_sheet_RH = df_sheet_RH[['订单来源','业务类型','下单时间','商品名称','发展渠道编码','发展渠道名称','发展人编码','发展人名称','发展渠道所属分公司']]  

    new_columns = ['订单来源','业务类型','下单时间','商品名称','渠道编码','渠道名称','发展人编码','发展人名称','下单分公司']
    df_sheet_YWWL.columns =  new_columns
    df_sheet_YWXC.columns =  new_columns
    df_sheet_LLB.columns =  new_columns
    df_sheet_DK.columns =  new_columns
    #df_sheet_RH.columns =  new_columns
    
    df = pd.concat([df_sheet_YWWL,df_sheet_YWXC], axis=0, ignore_index=True)
    df = pd.concat([df,df_sheet_LLB], axis=0, ignore_index=True)  
    df = pd.concat([df,df_sheet_DK], axis=0, ignore_index=True)  
    #df = pd.concat([df,df_sheet_RH], axis=0, ignore_index=True)    
    
    #20200408解决空值问题
    df1 = df.replace(np.nan, '空白', regex=True)
    
    #print(df)

        #写入文件夹
    writer = pd.ExcelWriter(file_reasoult)

    df1.to_excel(writer,sheet_name='订单明细汇总',index = False)


    writer.save()
    writer.close()
    
def all_deal_excel(file_in_daily,file_in_all,file_out):
    df_sheet_valid_daily = pd.DataFrame(pd.read_excel(file_in_daily,sheet_name='名单制渠道明细'))
    df_sheet_invalid_daily = pd.DataFrame(pd.read_excel(file_in_daily,sheet_name='非名单制渠道明细'))  
    df_sheet_valid_all = pd.DataFrame(pd.read_excel(file_in_all,sheet_name='名单制渠道订单明细'))
    df_sheet_invalid_all = pd.DataFrame(pd.read_excel(file_in_all,sheet_name='非名单制渠道订单明细'))   
    
    df_sheet_valid = pd.concat([df_sheet_valid_daily,df_sheet_valid_all], axis=0, ignore_index=True)
    df_sheet_invalid = pd.concat([df_sheet_invalid_daily,df_sheet_invalid_all], axis=0, ignore_index=True)    

    #2.按渠道编码计数+去重

    df_deal_reasoult_count = df_sheet_valid.groupby(['渠道编码','下单分公司']).size().sort_values(ascending=False)
    df_deal_reasoult_count = df_deal_reasoult_count.reset_index(drop = False) #修改索引
    df_deal_reasoult_count = df_deal_reasoult_count.rename(columns = {0:'计数'}) #规范名称
    #print('去重计数：')
    #print(df_deal_reasoult_count)    
    
    #3.求下单分公司数量
    df_deal_reasoult_count_company = df_deal_reasoult_count.groupby(['下单分公司']).size().sort_values(ascending=False)
    df_deal_reasoult_count_company = df_deal_reasoult_count_company.reset_index(drop = False) #修改索引
    df_deal_reasoult_count_company = df_deal_reasoult_count_company.rename(columns = {0:'计数'}) #规范名称
    
    
    
    
    #写入文件夹
    writer = pd.ExcelWriter(file_out)
    
    df_sheet_valid_daily.to_excel(writer,sheet_name='当月名单制渠道订单明细',index = False)
    df_sheet_invalid_daily.to_excel(writer,sheet_name='当月非名单制渠道订单明细',index = False)
    df_sheet_valid.to_excel(writer,sheet_name='总名单制渠道订单明细',index = False)
    df_sheet_invalid.to_excel(writer,sheet_name='总非名单制渠道订单明细',index = False)
    df_deal_reasoult_count_company.to_excel(writer,sheet_name='名单制渠道计数',index = False)

    writer.save()
    writer.close()
    
    
    
def main():
        

    
    #1.汇总原始数据
    #arr_gather_resoult = gather_excel(file,file_reasoult)
    #print(arr_gather_resoult)
    gather_excel_DF(file,file_reasoult)
    
    '''
    #array类型数据,用自己重写的MyRead_excel得到
    arr_gather_reasoult_read = MyRead_excel(file_reasoult,0)
    arr_rule = MyRead_excel(file_rule,0)
    '''
    
    #dataframe类型数据,用自带的read_excel得到
    df_gather_reasoult_read = pd.DataFrame(pd.read_excel(file_reasoult))
    df_rule = pd.DataFrame(pd.read_excel(file_rule))

    
    
    #print(df_gather_reasoult_read)
    #print(df_rule)
    
    

    #2.按渠道编码计数+去重

    df_gather_reasoult_count = df_gather_reasoult_read.groupby(['渠道编码','渠道名称','下单分公司']).size().sort_values(ascending=False)
    df_gather_reasoult_count = df_gather_reasoult_count.reset_index(drop = False) #修改索引
    df_gather_reasoult_count = df_gather_reasoult_count.rename(columns = {0:'计数'}) #规范名称
    print('去重计数：')
    print(df_gather_reasoult_count)

    #3.插值
    #df_gather_reasoult_read_deal = df_gather_reasoult_read[['渠道编码','渠道名称','下单分公司']] #取值
    #df_gather_reasoult_detail_all = pd.merge(df_gather_reasoult_count,df_gather_reasoult_read_deal,on = '渠道编码') #内联
    #df_gather_reasoult_detail_all = df_gather_reasoult_detail_all.drop_duplicates(subset=['渠道编码'])#去重
    #print(df_gather_reasoult_detail_all)

    df_gather_reasoult_detail_all = pd.merge(df_gather_reasoult_count,df_rule[['渠道编码','渠道属性']],how = 'left',on = ['渠道编码'])
    df_gather_reasoult_detail_all = df_gather_reasoult_detail_all[['下单分公司','渠道编码','渠道名称','渠道属性','计数']]
    print(df_gather_reasoult_detail_all)

     
    #4.分类
    df_gather_reasoult_detail_all = df_gather_reasoult_detail_all.fillna(value='非名单制渠道')    
    print('插值汇总：')
    print(df_gather_reasoult_detail_all)    
    df_gather_reasoult_detail_valid = df_gather_reasoult_detail_all[(df_gather_reasoult_detail_all['渠道属性'] != '非名单制渠道')]
    print('名单制：')
    print(df_gather_reasoult_detail_valid)
    df_gather_reasoult_detail_invalid = df_gather_reasoult_detail_all[(df_gather_reasoult_detail_all['渠道属性'] == '非名单制渠道')]
    df_gather_reasoult_detail_invalid = df_gather_reasoult_detail_invalid [['下单分公司','渠道编码','渠道名称','计数']]
    print('非名单制：')
    print(df_gather_reasoult_detail_invalid)
    
    #4.2处理计数问题
    df_gather_reasoult_detail_valid_count = df_gather_reasoult_detail_valid
    df_gather_reasoult_detail_valid_count = df_gather_reasoult_detail_valid_count.drop_duplicates(subset=['渠道编码'])#去重
    df_gather_reasoult_detail_valid_count =pd.DataFrame(df_gather_reasoult_detail_valid_count['下单分公司'].value_counts()) #计数
    df_gather_reasoult_detail_valid_count = df_gather_reasoult_detail_valid_count.reset_index(drop = False) #修改索引
    df_gather_reasoult_detail_valid_count = df_gather_reasoult_detail_valid_count.rename(columns = {'index':'下单分公司','下单分公司':'计数'}) #规范名称
    print('分公司计数：')
    print(df_gather_reasoult_detail_valid_count)

    #写入文件夹
    writer = pd.ExcelWriter(file_detail)

    df_gather_reasoult_detail_valid.to_excel(writer,sheet_name='名单制渠道明细',index = False)
    df_gather_reasoult_detail_invalid.to_excel(writer,sheet_name='非名单制渠道明细',index = False)
    df_gather_reasoult_detail_valid_count.to_excel(writer,sheet_name='名单制渠道计数',index = False)   

    writer.save()
    writer.close()

    
if __name__ == '__main__':
    '''
    处理步奏：
    1.汇总
    2.计数+去重
    3.插值
    4.分类
    '''
    
    path = 'V:\\工作\\联通\\数据\\20200426'
    path_report = 'V:\\工作\\联通\\数据\\通报'
    
    resourse = '4月25日分公司小程序明细.xlsx'
    reasoult = 'GatherReasoult.xlsx'
    rule =  '有效渠道码表-3月20日更新.xlsx'  
    detail = 'Detail.xlsx'
    detail_accumulative = '分公司小程序注册及业务发展统计表(3月).xlsx'
    detail_all = 'Detail_all.xlsx'

    file = path + '\\' + resourse #源文件地址
    file_reasoult = path + '\\' + reasoult   #生成新文件地址  
    file_rule = path_report + '\\' + rule
    file_detail = path + '\\' + detail
    file_detail_accumulative = path_report + '\\' + detail_accumulative
    file_detail_all = path + '\\' + detail_all    
    
    
    main()  #汇总+生成detail
    ##gather_excel_DF(file,file_reasoult)
    
    #处理detail生成Detail_all
    all_deal_excel(file_detail,file_detail_accumulative,file_detail_all)



